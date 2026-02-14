import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict

DATA_SRC = "./data-split-by-variable"
DATA_OUT = "./data"

os.makedirs(DATA_OUT, exist_ok=True)

def parse_filename(fname):
    """
    è§£ææª”åï¼Œä¾‹å¦‚ï¼š
    Denmark-2015A.xlsx
    Denmark-2015-2018B.xlsm
    å›å‚³ (country, year_start, year_end, variable_tag)
    """
    name = os.path.splitext(fname)[0]
    m = re.match(r"(.+?)-(\d{4})(?:-(\d{4}))?([A-Z]+)$", name)

    if not m:
        return None

    country, y1, y2, vars_ = m.groups()
    y2 = y2 if y2 else y1

    return [
        (country, int(y1), int(y2), var, fname)
        for var in vars_
    ]

def get_expected_output_files(parsed, country_year_spans):
    outputs = {}  # out_path -> (country, year_label)

    for country, spans in country_year_spans.items():
        is_consistent, year_span_list = check_year_span_consistency(
            country, spans
        )
        if not is_consistent:
            continue

        for start_year, end_year in year_span_list:
            year_label = (
                f"{start_year}"
                if start_year == end_year
                else f"{start_year}-{end_year}"
            )
            fname = f"{country}-{year_label}.xlsx"
            out_path = os.path.join(DATA_OUT, fname)
            outputs[out_path] = (country, year_label)

    return outputs

def create_output_file(country, start_year, end_year):
    year_label = (
        f"{start_year}"
        if start_year == end_year
        else f"{start_year}-{end_year}"
    )

    fname = f"{country}-{year_label}.xlsx"

    out_path = os.path.join(DATA_OUT, fname)
    
    files = [f for f in os.listdir(DATA_SRC) if f.endswith((".xlsx", ".xlsm"))]

    try:
        template_fname = find_excel_file(country, start_year, "A", files)
    except FileNotFoundError:
        raise FileNotFoundError(
            f"âŒ æ‰¾ä¸åˆ° {country}-{start_year}A.xlsx æˆ– .xlsm ä½œç‚ºæ¨¡æ¿"
        )

    template_path = os.path.join(DATA_SRC, template_fname)

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"æ‰¾ä¸åˆ°æª”æ¡ˆï¼š{template_path}")
    
    wb = load_workbook(template_path)

    if "REQUEST_TABLE" not in wb.sheetnames:
        raise ValueError(f"{template_fname} ä¸­æ²’æœ‰ REQUEST_TABLE å·¥ä½œè¡¨")
    
    wb.save(out_path)
    
    return out_path

def find_excel_file(country, start_year, var_tag, files):
    """
    æ‰¾å‡ºæŒ‡å®šåœ‹å®¶ã€å¹´ä»½ã€è®Šæ•¸çš„æª”æ¡ˆï¼ˆA/B/C...ï¼‰
    æ”¯æ´å–®å¹´æˆ–è·¨å¹´
    """
    # ç²¾ç¢ºåŒ¹é… country-startyear(-endyear)var_tag
    pattern = re.compile(
        rf"^{re.escape(country)}-{start_year}(?:-\d{{4}})?[A-Z]*{var_tag}[A-Z]*\.(xlsx|xlsm)$"
    )
    candidates = [f for f in files if pattern.match(f)]

    if not candidates:
        raise FileNotFoundError(
            f"âŒ æ‰¾ä¸åˆ° {country}-{start_year}{var_tag}.xlsx æˆ– .xlsm"
        )

    # å¦‚æœå‰›å¥½æœ‰å…©å€‹ï¼ˆç†è«–ä¸Šä¸æ‡‰è©²ï¼‰ï¼Œå„ªå…ˆç”¨ .xlsx
    candidates.sort(key=lambda x: x.endswith(".xlsm"))
    return candidates[0]


def check_year_span_consistency(country, year_spans):
    """
    1) å°‡æ‰€æœ‰å¹´ä»½æ¨™æº–åŒ–ï¼Œå–®ä¸€å¹´ -> (year, year)
    2) ä¾é–‹å§‹å¹´æ’åºï¼Œæ‰¾å‡ºå„å€‹ year_span
    3) ç¢ºä¿åŒä¸€ year_span çš„ A/B/C/... å¹´æ®µå®Œå…¨ä¸€è‡´
    4) å¾ŒçºŒ year_span ä¸èƒ½é‡ç–Šå…ˆå‰ year_span å¹´ä»½
    å›å‚³ï¼š
        - is_consistent: True/False
        - year_span_list: list of (start_year, end_year)
    """
    # æ¨™æº–åŒ–ï¼šåªæœ‰ä¸€å¹´ -> (year, year)
    normalized_year_span = [(s, s) if e is None else (s, e) for s, e in year_spans]

    # ä¾é–‹å§‹å¹´æ’åº
    normalized_year_span = sorted(normalized_year_span, key=lambda x: x[0])

    year_span_list = []
    current_start, current_end = normalized_year_span[0]

    for s, e in normalized_year_span[1:]:
        if s <= current_end:  # å±¬æ–¼åŒä¸€å€‹ year_span
            if (s, e) != (current_start, current_end):
                print(f"\nğŸš¨ {country}ï¼šåŒä¸€ year_span A/B/C å¹´æ®µä¸ä¸€è‡´")
                print(f"  Expectedï¼š{current_start}-{current_end}")
                print(f"  Foundï¼š{s}-{e}")
                return False, None
        else:   # æ–° year_span
            year_span_list.append((current_start, current_end))
            current_start, current_end = s, e

    year_span_list.append((current_start, current_end))  # åŠ æœ€å¾Œä¸€å€‹ year_span

    # æª¢æŸ¥ year_span ä¹‹é–“ä¸é‡ç–Š
    for i in range(1, len(year_span_list)):
        prev_s, prev_e = year_span_list[i-1]
        curr_s, curr_e = year_span_list[i]
        if curr_s <= prev_e:
            print(f"\nğŸš¨ {country}ï¼šèˆ‡å‰ä¸€å€‹ year_span é‡ç–Š")
            print(f"  å‰ä¸€å€‹ year_spanï¼š{prev_s}-{prev_e}")
            print(f"  ç•¶å‰ year_spanï¼š{curr_s}-{curr_e}")
            return False, None

    return True, year_span_list

def read_request_table(xls_path):
    """è®€å– REQUEST_TABLEï¼Œå›å‚³ dataframeï¼ˆrow=7 é–‹å§‹ï¼‰"""
    return pd.read_excel(
        xls_path, sheet_name="REQUEST_TABLE", engine="openpyxl", header=None
    )

def get_sheet_for_year(req_df, year):
    """æ ¹æ“š REQUEST_TABLE æ‰¾åˆ°å°æ‡‰å¹´ä»½çš„å·¥ä½œè¡¨ä½ç½®"""
    
    # å¾ row7 é–‹å§‹æŠ“ Gæ¬„ï¼ˆindex=6ï¼‰
    df_years = pd.to_numeric(req_df.iloc[6:, 6], errors='coerce')
    matches = df_years[df_years == year]

    if matches.empty:
        print("ğŸ” DEBUGï¼šREQUEST_TABLE G æ¬„ 'Start Date'ï¼ˆå‰ 5 ç­†ï¼‰å…§å®¹å¦‚ä¸‹ï¼š")
        print(df_years.head(5).tolist())
        raise ValueError(f"âš ï¸ REQUEST_TABLE æ‰¾ä¸åˆ°å¹´ä»½ {year}")

    # å–ç¬¬ä¸€å€‹ç¬¦åˆå¹´ä»½çš„åˆ—ç´¢å¼•
    row_idx = matches.index[0]
    row_series = req_df.iloc[row_idx]

    sheet_ref = row_series[10]      # Kæ¬„
    expected_rows = row_series[13]  # Næ¬„
    expected_cols = row_series[14]  # Oæ¬„

    # sheet_ref å½¢å¦‚: å·¥ä½œè¡¨1'!$A$1
    sheet_name = sheet_ref.split("!") [0].replace("'", "")

    return sheet_name, int(expected_rows), int(expected_cols), row_idx + 1

def read_variable_data(xls_path, sheet_name):
    """å¾æŒ‡å®š sheet è®€è³‡æ–™"""
    df = pd.read_excel(xls_path, sheet_name=sheet_name, engine="openpyxl")
    return df

def append_column(out_path, df, sheet_name, variable_suffix):
    """
    ä»¥ A æ¬„ Type ç•¶ primary key åˆä½µ
    """

    wb = load_workbook(out_path)

    # è®€å–ç¾æœ‰ sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        data = ws.values
        columns = next(data)
        base_df = pd.DataFrame(data, columns=columns)

    else:
        # å¦‚æœä¸å­˜åœ¨ï¼Œç›´æ¥å¯«å…¥
        base_df = pd.DataFrame()

    # æ•´ç†æ–°è³‡æ–™
    df = df.copy()
    df.columns = df.columns.astype(str)

    if "Type" not in df.columns:
        raise ValueError("âŒ æ–°è³‡æ–™æ²’æœ‰ Type æ¬„")
    
    df["Type"] = df["Type"].astype(str).str.strip()

    # å¦‚æœ base æ˜¯ç©º
    if base_df.empty:
        merged_df = df

    else:
        if "Type" not in base_df.columns:
            raise ValueError("âŒ æ—¢æœ‰è³‡æ–™æ²’æœ‰ Type æ¬„")
        
        base_df["Type"] = base_df["Type"].astype(str).str.strip()
        base_df = base_df.dropna(subset=["Type"])

        
        # æ¸…æ‰æ®˜ç•™æ¬„ä½
        if "_order" in base_df.columns:
            base_df = base_df.drop(columns="_order")

        # ä¿ç•™åŸé †åº
        base_df["_order"] = range(len(base_df))

        # æ‹¿æ–°è³‡æ–™ã€Œé™¤äº† Type ä»¥å¤–ã€çš„æ¬„
        new_cols = [c for c in df.columns if c != "Type"]

        # ========= å…¬å¸å·®ç•°åˆ†æ =========
        base_types = list(base_df["Type"])
        new_types = list(df["Type"])

        base_index_map = {t: i+2 for i, t in enumerate(base_types)}  # +2 å› ç‚º Excel æœ‰è¡¨é ­
        new_index_map = {t: i+2 for i, t in enumerate(new_types)}

        set_base = set(base_types)
        set_new = set(new_types)

        only_in_new = sorted(set_new - set_base)
        only_in_base = sorted(set_base - set_new)

        merged_df = pd.merge(
            base_df,
            df[["Type"] + new_cols],
            on="Type",
            how="outer",
            sort=False
        )

        # æ’åºï¼šA åŸé †åºåœ¨å‰ï¼Œæ–°å…¬å¸æ’å¾Œ
        merged_df = merged_df.sort_values("_order", na_position="last")
        merged_df = merged_df.drop(columns=["_order"])

        # é‡æ–°å»ºç«‹ index map
        final_index_map = {
            t: i+2 for i, t in enumerate(merged_df["Type"])
        }

        # -------- æ–°å…¬å¸ --------
        for idx, company in enumerate(only_in_new):
            new_row_position = len(base_types) + idx + 2
            print(
                f"æ–°å…¬å¸ {company} å‡ºç¾åœ¨ {sheet_name}{variable_suffix} çš„ç¬¬ {new_index_map[company]} åˆ—ï¼Œ"
                f"åŠ é€² {sheet_name}A çš„ç¬¬ {final_index_map[company]} åˆ—"
            )

        # -------- å°‘å…¬å¸ --------
        for company in only_in_base:
            print(
                f"å…¬å¸ {company} å‡ºç¾åœ¨ {sheet_name}A çš„ç¬¬ {base_index_map[company]} åˆ—ï¼Œ"
                f"ä½†æ²’æœ‰å‡ºç¾åœ¨ {sheet_name}{variable_suffix}ï¼Œ"
                f"è©²å…¬å¸ {variable_suffix} çµ„è®Šæ•¸çš„å€¼å…¨éƒ¨è£œ ."
            )

    # å°‡ NaN è½‰ç‚º "."
    merged_df = merged_df.fillna(".")

    # æ¸…ç©ºèˆŠ sheet
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(title=sheet_name)

    # å¯«å›
    for r in dataframe_to_rows(merged_df, index=False, header=True):
        ws.append(r)

    wb.save(out_path)

def update_request_table(out_path, src_path, excel_row):
    """
    å°‡ src_path çš„ REQUEST_TABLE ä¸­
    å°æ‡‰ year çš„ O/P æ¬„ï¼ŒåŠ åˆ° out_path çš„ REQUEST_TABLE
    ä¸¦å°å‡ºåŠ ç¸½éç¨‹
    """
    wb_out = load_workbook(out_path)
    wb_src = load_workbook(src_path)

    ws_out = wb_out["REQUEST_TABLE"]
    ws_src = wb_src["REQUEST_TABLE"]

    # ========= å…ˆæª¢æŸ¥ N æ¬„ (Rows) =========
    N_COL = 14  # column N

    n_out = ws_out.cell(row=excel_row, column=N_COL).value
    n_src = ws_src.cell(row=excel_row, column=N_COL).value

    if n_out != n_src:
        print(
            f"âš ï¸ ROWS ä¸ä¸€è‡´ | "
            f"{os.path.basename(out_path)} N{excel_row}={n_out} | "
            f"{os.path.basename(src_path)} N{excel_row}={n_src}"
        )

    for col, label in [(15, "O"), (16, "P")]:
        v_out = ws_out.cell(row=excel_row, column=col).value or 0
        v_src = ws_src.cell(row=excel_row, column=col).value or 0
        new_v = v_out + v_src

        print(
            f"ğŸ§® {os.path.basename(out_path)} "
            f"REQUEST_TABLE {label}{excel_row}: "
            f"{v_out} + {v_src} = {new_v}"
        )

        ws_out.cell(row=excel_row, column=col, value=new_v)

    wb_out.save(out_path)

def main():
    files = [f for f in os.listdir(DATA_SRC) if f.endswith((".xlsx", ".xlsm"))]

    parsed = []
    for f in files:
        parsed.extend(parse_filename(f))

    # ä¾åœ‹å®¶ -> å¹´åº¦ -> è®Šæ•¸æ’åºï¼ˆA, B, C...ï¼‰
    grouped = defaultdict(lambda: defaultdict(list))  # country -> year -> list of (var, fname)
    country_year_spans = defaultdict(list)
    
    for country, y1, y2, var, fname in parsed:
        country_year_spans[country].append((y1, y2))
        for y in range(y1, y2 + 1):
            grouped[country][y].append((var, fname))
    
    # æª¢æŸ¥ä¹‹å‰æ˜¯å¦å·²è¼¸å‡ºé
    expected_outputs = get_expected_output_files(parsed, country_year_spans)

    existing_outputs = {
        path: meta
        for path, meta in expected_outputs.items()
        if os.path.exists(path)
    }

    if existing_outputs:
        print("\nâš ï¸ ç™¼ç¾ä»¥ä¸‹è¼¸å‡ºæª”å·²å­˜åœ¨ ./dataï¼š")
        for i, (path, (country, year_label)) in enumerate(existing_outputs.items(), 1):
            print(f"{i}. {country} ({year_label}) â†’ {os.path.basename(path)}")

        while True:
            ans = input(
                "\nğŸ‘‰ æ˜¯å¦ã€å…¨éƒ¨åˆªé™¤ã€‘å¾Œé‡æ–°ç”¢ç”Ÿï¼Ÿ (y/n): "
            ).strip().lower()

            if ans == "y":
                for path in existing_outputs:
                    print(f"ğŸ—‘ï¸ åˆªé™¤ {os.path.basename(path)}")
                    os.remove(path)
                break

            elif ans == "n":
                print(
                    "\nâ­ï¸ æœªåˆªé™¤ä»»ä½•æª”æ¡ˆã€‚\n"
                    "è«‹è‡ªè¡Œè‡³ ./data åˆªé™¤æ¬²é‡æ–°ç”¢ç”Ÿçš„æª”æ¡ˆå¾Œå†åŸ·è¡Œã€‚"
                )
                return

            else:
                print("è«‹è¼¸å…¥ y æˆ– n")

    for country, spans in grouped.items():

        # å…ˆæª¢æŸ¥è©²åœ‹æ‰€æœ‰æª”æ¡ˆçš„å¹´æ®µæ˜¯å¦ä¸€è‡´
        is_consistent, year_span_list = check_year_span_consistency(
            country, country_year_spans[country]
        )
        if not is_consistent:
            continue   # æ•´å€‹åœ‹å®¶ç›´æ¥è·³éï¼Œä¸è¼¸å‡º

        print(f"\n========== â–¶ é–‹å§‹è™•ç† {country} ==========")

        for start_year, end_year in year_span_list:
            print("\n" + "-" * 40)
            out_xlsx = create_output_file(country, start_year, end_year)
            if out_xlsx is None:
                continue   # é€™å€‹å¹´åº¦å·²åšéï¼Œç›´æ¥è·³é
            skip_country = False

            # ç¯©é¸é€™å€‹ block çš„æª”æ¡ˆ
            block_files = [
                (y1, y2, var, fname)
                for parsed_country, y1, y2, var, fname in parsed
                if parsed_country == country and y1 >= start_year and y2 <= end_year
            ]
            block_files = sorted(block_files, key=lambda x: x[2])  # A/B/C æ’åºï¼Œä»¥ç¬¬ä¸€å€‹æœ€å°å­—æ¯å…ˆè™•ç†
            
            processed_files = set() # è¨˜éŒ„å·²è™•ç† Excel

            for s, e, var, _ in block_files:
                fname = find_excel_file(country, s, var, files)
                if fname in processed_files:
                    continue  # å¦å‰‡ Hong-Kong-2015CD æœƒè¢«ä½µ 2 æ¬¡
                processed_files.add(fname)  # æ¨™è¨˜ Hong-Kong-2015CD å·²è™•ç†

                src_path = os.path.join(DATA_SRC, fname)
                vars_in_file = [v for _, _, v, f in block_files if f == fname]
                is_first_variable = ("A" in vars_in_file)
                print(f"ğŸ“‚ è™•ç† {src_path}")

                req_df = read_request_table(src_path)

                for year in range(s, e+1):
                    try:
                        sheet_name, exp_rows, exp_cols, excel_row = get_sheet_for_year(req_df, year)
                        df = read_variable_data(src_path, sheet_name)
                        df_rows, df_cols = df.shape  # DataFrame ä¸å« headerï¼Œæœƒå°‘ä¸€ row

                        actual_rows = df_rows + 1
                        actual_cols = df_cols

                        # æª¢æŸ¥å°ºå¯¸
                        if actual_rows != exp_rows or actual_cols != exp_cols:
                            print(f"âš ï¸ {country}-{start_year}-{end_year}{var} rows/cols ä¸ç¬¦"
                                f"   Expected: {exp_rows} rows x {exp_cols} cols\n"
                                f"   Actual:   {actual_rows} rows x {actual_cols} cols"
                            )
                            skip_country = True
                            break
                        else:
                            print(f"ğŸ”¹ å·¥ä½œè¡¨: {sheet_name}, shape: {exp_rows} rows x {exp_cols} columns")

                        if is_first_variable:   # A çµ„è®Šæ•¸ä½œç‚ºæ¨¡æ¿ï¼Œå·²ç¶“åœ¨æ–°æª”è£¡ï¼Œskip
                            continue

                        append_column(
                            out_path=out_xlsx,
                            df=df,
                            sheet_name=sheet_name,
                            variable_suffix=var
                        )

                        if not is_first_variable:
                            update_request_table(
                                out_path=out_xlsx,
                                src_path=src_path,
                                excel_row=excel_row
                            )
                    except Exception as e:
                        print(f"âš ï¸ ERROR: {e}")
                        skip_country = True
                        break   # è·³å‡º var è¿´åœˆï¼Œå¤–å±¤æœƒè™•ç†åˆªæª” + æ›åœ‹

            if skip_country:
                if os.path.exists(out_xlsx):
                    print(f"ğŸ—‘ï¸ åˆªé™¤æª”æ¡ˆ {out_xlsx}")
                    os.remove(out_xlsx)
                break   # è·³å‡º year è¿´åœˆ (ç•¥éå¾ŒçºŒå¹´åº¦)ï¼Œæ›ä¸‹ä¸€åœ‹

    print("ğŸ‰ æ‰€æœ‰åœ‹å®¶/å¹´åº¦æ•´åˆå®Œæˆï¼")

if __name__ == "__main__":
    main()

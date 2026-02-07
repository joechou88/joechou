import re
import os
from openpyxl import load_workbook

# ================== 設定 ==================
INPUT_FOLDER = "data-split-by-entity"
OUTPUT_FOLDER = "data-split-by-variable"
REQUEST_SHEET = "REQUEST_TABLE"

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# ================== 檔名解析 ==================
pattern = re.compile(
    r"""
    (?P<country>[A-Za-z]+)
    (?P<company>\d+)
    -
    (?P<start>\d{4})
    (?:-(?P<end>\d{4}))?
    (?P<suffix>[A-Za-z]+)
    """,
    re.VERBOSE
)

def parse_filename(fname):
    name = os.path.splitext(fname)[0]
    m = pattern.fullmatch(name)
    return m.groupdict() if m else None

def check_request_table(wb, fname, company_no, start, end):
    ws = wb[REQUEST_SHEET]
    expected_series = f"FDEALL{company_no}"

    start_year = int(start)
    end_year = int(end) if end else start_year
    expected_years = list(range(start_year, end_year + 1))

    row = 7
    year_idx = 0

    while ws[f"E{row}"].value not in (None, ""):
        # ===== E 欄：公司組數檢查 =====
        if ws[f"E{row}"].value != expected_series:
            print(
                f"⚠️ 跳過: {fname} REQUEST_TABLE E{row} = {ws[f'E{row}'].value}，"
                f"預期 {expected_series}"
            )
            return False

        # ===== G 欄：年份檢查 =====
        if year_idx >= len(expected_years):
            print(
                f"⚠️ 跳過: {fname} REQUEST_TABLE 年份列數超出檔名範圍（從 G{row} 開始）"
            )
            return False

        raw_year = ws[f"G{row}"].value
        expected_year = expected_years[year_idx]

        try:
            cell_year = int(str(raw_year).strip())
        except Exception:
            print(
                f"⚠️ 跳過: {fname} REQUEST_TABLE G{row} = {raw_year}，"
                f"無法解析為年份"
            )
            return False

        if cell_year != expected_year:
            print(
                f"⚠️ 跳過: {fname} REQUEST_TABLE G{row} = {cell_year}，"
                f"預期 {expected_year}（與檔名年份不一致）"
            )
            return False

        row += 1
        year_idx += 1

    # ===== 列數反向檢查（避免少一年）=====
    if year_idx != len(expected_years):
        raise ValueError(
            f"{fname} REQUEST_TABLE 年份列數不足，"
            f"預期 {len(expected_years)} 列，實際 {year_idx} 列"
        )

def validate_wb(wb, fname, company_no, start, end):
    # ===== 確定 REQUEST_TABLE 存在 =====
    if REQUEST_SHEET not in wb.sheetnames:
        raise ValueError(f"{fname} 缺少 REQUEST_TABLE")
    
    # ===== 檢查 檔名和 REQUEST_TABLE 的 Series 一致 =====
    check_request_table(wb, fname, company_no, start, end)
    
    # ===== 檢查 檔名和 工作表數量 一致 =====
    data_sheets = [s for s in wb.sheetnames if s != REQUEST_SHEET]
    if len(data_sheets) < years:
        raise ValueError(
            f"{fname} 工作表數量不足，預期 {years} 張，實際 {len(data_sheets)}"
        )

def print_sheet_shapes(wb, fname, skip_sheet=REQUEST_SHEET):
    """
    印出 workbook 每個 sheet 的 shape
    - wb: Workbook 物件
    - title: log 標題
    - skip_sheet: 不印的 sheet 名稱（預設 REQUEST_TABLE）
    """
    for ws_name in wb.sheetnames:
        if ws_name == skip_sheet:
            continue
        ws = wb[ws_name]
        rows = actual_rows(ws)
        cols = actual_cols(ws)
        print(f"{fname} 🔹 工作表: {ws_name}, "
              f"shape: {rows} rows x {cols} columns")

def actual_rows(ws):
    """
    計算實際有資料的 row 數（忽略尾端空白列）
    """
    last = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if any(cell is not None for cell in row):
            last = i
    return last

def actual_cols(ws):
    """
    計算實際有資料的欄位數（忽略尾端空欄）
    """
    max_cols = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        # 找最後一個非 None 的 index
        for i in range(len(row), 0, -1):
            if row[i-1] is not None:
                max_cols = max(max_cols, i)
                break
    return max_cols

# ================== row append ==================
def append_sheet_rows(target_ws, source_ws, fname_only):
    """
    將 source_ws 的資料接到 target_ws 後面
    - 只允許欄位數一致
    - 不一致時印出警告，但仍跳過 append
    """
    target_cols = actual_cols(target_ws)
    source_cols = actual_cols(source_ws)

    if target_cols != source_cols:
        print(f"⚠️ 跳過：{fname_only} 工作表 {source_ws.title} 欄位數不一致！"
              f"target: {target_cols}, source: {source_cols}")
        return  # 不 append

    for row in source_ws.iter_rows(min_row=2, values_only=True):
        target_ws.append(row)

# ================== 主流程 ==================
try:
    expected_company_count = int(
        input("🧩 請輸入每個國家預期的公司群數（例如 8）: ").strip()
    )
    if expected_company_count < 1:
        raise ValueError
except ValueError:
    print("❌ 請輸入大於等於 1 的整數")
    exit(1)

files = [
    f for f in os.listdir(INPUT_FOLDER)
    if f.endswith((".xlsx", ".xlsm"))
]

groups = {}
key_to_outname = {}

for f in files:
    info = parse_filename(f)
    if not info:
        continue
    key = (
        info["country"],
        info["start"],
        info["end"],
        info["suffix"]
    )
    groups.setdefault(key, []).append((int(info["company"]), f))

    if key not in key_to_outname:
        out_name = f"{info['country']}-{info['start']}{'-'+info['end'] if info['end'] else ''}{info['suffix']}.xlsx"
        key_to_outname[key] = out_name

missing_company_report = []
existing_outputs = []

for (country, start, end, suffix) in groups.keys():
    out_name = key_to_outname[(country, start, end, suffix)]
    out_path = os.path.join(OUTPUT_FOLDER, out_name)

    if os.path.exists(out_path):
        existing_outputs.append(out_path)

if existing_outputs:
    print("\n⚠️  以下輸出檔案已存在，將被覆蓋：")
    for p in existing_outputs:
        print(f"   - {p}")

    ans = input("\n是否同意刪除並全部重生？(y/N): ").strip().lower()

    if ans not in ("y", "yes"):
        print(
            "\n❌ 已取消執行。\n"
            "請自行到 ./data-split-by-variable 刪除上述檔案後再重新執行。"
        )
        exit(1)

    for p in existing_outputs:
        os.remove(p)
        print(f"🗑 已刪除：{p}")
    print(f"\n========================\n")

for (country, start, end, suffix), items in groups.items():    
    companies = {company: fname for company, fname in items}
    actual_companies = set(companies.keys())
    expected_companies = set(range(1, expected_company_count + 1))
    missing_companies = sorted(expected_companies - actual_companies)

    if missing_companies:
        missing_company_report.append({
            "country": country,
            "period": f"{start}{'-' + end if end else ''}{suffix}",
            "missing": missing_companies
        })

    # ===== 嚴格檢查：一定要有 company = 1 作為模板 =====
    if 1 not in companies:
        raise ValueError(
            f"缺少 company=1，無法合併：{country}-{start}{'-'+end if end else ''}{suffix}"
        )

    base_company = 1
    base_file = os.path.join(INPUT_FOLDER, companies[1])

    wb_base = load_workbook(base_file, data_only=True)
    years = 1 if end is None else int(end) - int(start) + 1

    validate_wb(wb_base, base_file, base_company, start, end)
    print_sheet_shapes(wb_base, companies[1])

    for company in sorted(companies):
        if company == 1:
            continue
        fname_only = companies[company]
        fname = os.path.join(INPUT_FOLDER, fname_only)
        wb_src = load_workbook(fname, data_only=True)

        validate_wb(wb_src, fname, company, start, end)

        for ws_name in wb_base.sheetnames:
            # 跳過 REQUEST_TABLE
            if ws_name == REQUEST_SHEET:
                continue

            ws_base = wb_base[ws_name]
            ws_src = wb_src[ws_name]

            rows = actual_rows(ws_src)
            cols = actual_cols(ws_src)

            print(
                f"{fname_only} 🔹 工作表: {ws_name}, "
                f"shape: {rows} rows x {cols} columns"
            )
            append_sheet_rows(ws_base, ws_src, fname_only)

    out_name = key_to_outname[(country, start, end, suffix)]
    out_path = os.path.join(OUTPUT_FOLDER, out_name)

    print(f"\n📊 {out_name} 最終合併後 sheet shape：")
    print_sheet_shapes(wb_base, out_name)
                       
    wb_base.save(out_path)
    print(f"✔ 輸出完成：{out_path}")
    print(f"\n========================\n")

if missing_company_report:
    print("\n⚠️ 公司群數量警示（不影響輸出）")
    print("====================================")
    for item in missing_company_report:
        print(
            f"{item['country']}-{item['period']} "
            f"缺少公司群：{', '.join(map(str, item['missing']))}"
        )
else:
    print("\n✅ 所有國家公司群數量皆符合預期")

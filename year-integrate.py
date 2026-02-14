import os
import pandas as pd
import re
import sys
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook, Workbook

# ========= 基本設定 =========
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_DIR = os.path.join(BASE_DIR, "data")
OUT_DIR = os.path.join(BASE_DIR, "data-2015-2024")
LOG_FILE = f"year_integrate_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

START_YEAR = 2015
END_YEAR = 2024

os.makedirs(OUT_DIR, exist_ok=True)

# 同時印到終端機 + log 檔案
class Tee:
    def __init__(self, *files):
        self.files = files
    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush()
    def flush(self):
        for f in self.files:
            f.flush()

def parse_years_from_filename(filename):
    """
    從檔名擷取合法年份集合
    Denmark-2014.xlsx -> {2014}
    Denmark-2015-2017A.xlsm -> {2015, 2016, 2017}
    """
    name = os.path.splitext(filename)[0]

    m = re.search(r"-(\d{4})(?:-(\d{4}))?", name)
    if not m:
        return set()

    start = int(m.group(1))
    end = int(m.group(2)) if m.group(2) else start

    return set(range(start, end + 1))

def parse_country(filename):
    """從檔名擷取國家名稱"""
    name = os.path.splitext(filename)[0]
    country = re.sub(r"-\d{4}(-\d{4})?$", "", name)
    return country

def extract_sheet_name(ref):
    """從 K 欄 'Sheet1'!$A$1 抽出 Sheet1"""
    return ref.split("!")[0].replace("'", "")

def main():
    # ========= 收集各國檔案 =========
    country_files = defaultdict(list)

    for f in os.listdir(SRC_DIR):
        if f.lower().endswith((".xlsm", ".xlsx")):
            country_files[parse_country(f)].append(f)

    # ---- 讀 country code 對照表 ----
    code_df = pd.read_excel(os.path.join(BASE_DIR, "country-code.xlsx"))
    code_df["Country_name"] = code_df["Country_name"].str.strip()

    country_code_map = code_df.set_index("Country_name").to_dict(orient="index")

    # ========= 主流程 =========
    for country, files in country_files.items():
        print(f"\nProcessing {country}...")

        display_country = country.replace("-", " ") # 取得 COUNTRY 欄

        records = []            # 暫存 某個國家 各年度的資料
        header = None
        expected_cols = None    # 紀錄該國家應有的欄位數
        year_col_count = {}     # 紀錄每年欄位數（方便報錯）

        # ---- 掃描該國所有來源檔 ----
        for fname in files:
            path = os.path.join(SRC_DIR, fname)
            wb = load_workbook(path, data_only=True)

            if "REQUEST_TABLE" not in wb.sheetnames:
                wb.close()
                print(f"❌ 缺少 REQUEST_TABLE！略過國家：{country}")
                continue

            req_ws = wb["REQUEST_TABLE"]
            row = 7

            # REQUEST_TABLE 從第 7 列一直往下讀，讀到空白就停
            while True:
                file_years = parse_years_from_filename(fname)

                year = req_ws[f"G{row}"].value
                ref = req_ws[f"K{row}"].value

                if year is None or ref is None:
                    break

                year = int(year)

                # ====== 檔名 vs REQUEST_TABLE 年份檢查 ======
                if year not in file_years:
                    print(
                        f"❌ 年份不一致｜檔名: {fname} "
                        f"| REQUEST_TABLE 年份: {year} "
                        f"| 檔名年份: {sorted(file_years)} → 已跳過"
                    )
                    row += 1
                    continue

                if START_YEAR <= year <= END_YEAR:
                    sheet_name = extract_sheet_name(ref)

                    if sheet_name in wb.sheetnames:
                        src_ws = wb[sheet_name]
                        raw_rows = list(src_ws.iter_rows(values_only=True))
                        rows = [
                            r for r in raw_rows
                            if any(cell is not None for cell in r) # Excel 被更動過會殘留「看不見的空白列」，需自動丟棄
                        ]

                        # 若有丟棄空白列，印警告
                        if len(rows) < len(raw_rows):
                            print(
                                f"⚠ 警告｜{country} {year} 年："
                                f"工作表包含 {len(raw_rows)-len(rows)} 列殘留空白列，已自動移除"
                            )

                        if not rows:
                            row += 1
                            continue

                        # ====== 確保同一國家變數數量都一樣 ======
                        # 優先檢查 REQUEST_TABLE O 欄
                        # 備援：實際去數後面工作表欄位 - 1
                        cols_value = req_ws[f"O{row}"].value
                        rows_value = req_ws[f"N{row}"].value
                        print(
                            f"國家: {country} | 年份: {year} "
                            f"| O欄(cols_value) = {cols_value} "
                            f"| N欄(rows_value) = {rows_value}"
                            f"| 工作表列數={len(rows)}"
                        )

                        if isinstance(cols_value, int):
                            n_cols = cols_value
                        else:
                            n_cols = src_ws.max_column
                        number_of_variables = n_cols - 1 # 排除 Type (DSCD)

                        year_col_count[year] = n_cols

                        if expected_cols is None:
                            expected_cols = number_of_variables # 紀錄該國第一年變數數量
                        else:
                            if number_of_variables != expected_cols:
                                print(f"❌ 變數數量不一致，已略過該年份！國家：{country} 年份：{year}")
                                print(f"  期望變數數量（不含 Type/DSCD）：{expected_cols}")
                                print(f"  年份 {year} 變數數量：{number_of_variables}")
                                print("  各年變數數量（不含 Type/DSCD）：")
                                for y, c in year_col_count.items():
                                    print(f"   - {y}: {c-1}")
                                row += 1
                                continue # 跳回 while True 的開頭，跑下一年
                        # ====== 檢查結束 ======

                        if header is None:
                            header = ["YEAR"] + list(rows[0]) # 第一次跑該國時，紀錄欄位名稱

                        for data_row in rows[1:]:
                            records.append([year] + list(data_row)) # 把「某一年的資料」一筆一筆加進 MASTER_TABLE

                row += 1

            wb.close()

        if len(records) == 0:
            print(f"  ⚠ {country} 無有效資料，略過")
            continue

        # ---- 加 COUNTRY / COUNTRY_CODE / COUNTRY_CODE2 ----
        code_info = country_code_map.get(display_country, {"Country_code": "", "Country_code2": ""})
        country_code = code_info.get("Country_code", "")
        country_code2 = code_info.get("Country_code2", "")

        # 調整 header，把三欄插到 YEAR 之後
        new_header = header[:1] + ["COUNTRY", "COUNTRY_CODE", "COUNTRY_CODE2"] + header[1:]

        # 調整每筆資料
        new_records = []
        for row in records:
            new_row = row[:1] + [display_country, country_code, country_code2] + row[1:]
            new_records.append(new_row)

        # ---- 依年份升冪排序 ----
        new_records.sort(key=lambda x: x[0])

        # ---- 取得實際年份範圍 ----
        years_present = sorted({row[0] for row in new_records})

        # ---- 檢查年份完整性 ----
        required_years = set(range(START_YEAR, END_YEAR + 1))
        missing_required = sorted(required_years - set(years_present))

        # ---- 檢查年份連續性 ----
        min_year = min(years_present)
        max_year = max(years_present)
        full_range = set(range(min_year, max_year + 1))
        missing_continuous = sorted(full_range - set(years_present))

        # ---- 若任一條件不符合就跳過 ----
        if missing_required or missing_continuous:
            msg = f"⚠ {country} 資料不完整或不連續｜實際年份: {years_present}"
            if missing_required:
                msg += f" | 年份不完整，缺少: {', '.join(str(y) for y in missing_required)}"
            if missing_continuous:
                msg += f" | 年份不連續，缺少: {', '.join(str(y) for y in missing_continuous)}"
            print(msg)
            continue  # 跳過該國家，不輸出

        # ---- 輸出主控表 ----
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "MASTER_TABLE"

        out_ws.append(new_header)
        for r in new_records:
            out_ws.append(r)

        out_path = os.path.join(
            OUT_DIR, f"{country}-{min_year}-{max_year}.xlsx"
        )
        out_wb.save(out_path)

        print(f"  ✔ 輸出完成: {out_path}，共 {len(new_records)} 筆資料")

    print("=== 全部國家彙整完成 ===")

if __name__ == "__main__":
    # 開啟 log（每次覆寫；若想改成累加，用 "a"）
    log_f = open(LOG_FILE, "w", encoding="utf-8")

    sys.stdout = Tee(sys.stdout, log_f)
    sys.stderr = Tee(sys.stderr, log_f)   # 錯誤也寫入 log
    
    print("="*60)
    print("Year Integration Log")
    print("Start Time:", datetime.now())
    print("="*60)

    try:
        main()
    except Exception as e:
        import traceback
        print("\n❌ 系統發生未預期錯誤：")
        traceback.print_exc()
    finally:
        print("\nEnd Time:", datetime.now())
        print("="*60)
        log_f.close()

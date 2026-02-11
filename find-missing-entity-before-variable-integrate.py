import os
import re
from collections import defaultdict
from openpyxl import load_workbook

# ==============================
# 設定
# ==============================

DATA_DIR = "./data-split-by-variable-all"
TYPE_COL_INDEX = 0      # A 欄 = Type
START_ROW = 2           # 第 1 列是表頭
IGNORE_SHEETS = {"REQUEST_TABLE"}

# ==============================
# 工具函式
# ==============================

def parse_filename(filename):
    """
    支援：
    South-Korea-2015A.xlsx
    South-Korea-2015D.xlsm
    Switzerland-2015-2018B.xlsx
    """
    m = re.match(
        r"(.+?-\d{4}(?:-\d{4})?)([A-Z]+)\.(xlsx|xlsm)$",
        filename,
        re.IGNORECASE
    )
    if not m:
        return None
    return m.group(1), m.group(2)


def read_excel_types(filepath):
    """
    回傳：
    {
        sheet_name: {
            Type: row_index
        }
    }
    （自動略過 REQUEST_TABLE）
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    result = {}

    for sheet in wb.worksheets:
        if sheet.title in IGNORE_SHEETS:
            continue

        type_map = {}
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=START_ROW, values_only=True),
            start=START_ROW
        ):
            type_val = row[TYPE_COL_INDEX]
            if type_val:
                type_map[str(type_val).strip()] = row_idx

        result[sheet.title] = type_map

    return result


# ==============================
# 主流程
# ==============================

def main():
    files = os.listdir(DATA_DIR)

    # group_key -> variable_group -> sheet -> {Type: row}
    data = defaultdict(dict)

    for f in files:
        parsed = parse_filename(f)
        if not parsed:
            continue

        group_key, var_group = parsed
        path = os.path.join(DATA_DIR, f)

        print(f"📂 讀取 {f}")
        data[group_key][var_group] = read_excel_types(path)

    print("\n================ 比對結果 ================\n")

    for group_key, group_data in data.items():
        print(f"🔍 檢查 {group_key}")

        var_groups = sorted(group_data.keys())
        sheet_names = group_data[var_groups[0]].keys()

        for sheet in sheet_names:
            # 所有變數組的 Type 聯集
            all_types = set()
            for g in var_groups:
                all_types |= set(group_data[g][sheet].keys())

            for g in var_groups:
                current_types = set(group_data[g][sheet].keys())
                missing = all_types - current_types

                for t in sorted(missing):
                    exists_in = {
                        other_g: group_data[other_g][sheet][t]
                        for other_g in var_groups
                        if other_g != g and t in group_data[other_g][sheet]
                    }

                    print(
                        f"  [{sheet}] "
                        f"{g} 少了 {t} ｜"
                        f"存在於 {exists_in}"
                    )

        print("-" * 50)


if __name__ == "__main__":
    main()
